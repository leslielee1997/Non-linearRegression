from scipy.optimize import minimize
import numpy as np
import matplotlib.pyplot as plt
import openpyxl as xl
plt.rcParams['font.sans-serif']=['SimHei'] 
plt.rcParams['axes.unicode_minus']=False   
Filename = 'F:\code\meachine-learning\data.xlsx'
Sheet = 'Sheet1'
data = xl.load_workbook(Filename)[Sheet]
def getColValues(data,column):
    rows = data.max_row
    columndata=[]
    for i in range(1,rows+1):
        cellvalue = data.cell(row=i,column=column).value
        columndata.append(cellvalue)
    return columndata
data_x = np.array(getColValues(data,1))
data_y = np.array(getColValues(data,2))
data_n = np.array(getColValues(data,3))

class GPR:

    def __init__(self, optimize=True):
        self.is_fit = False
        self.train_X, self.train_y = None, None
        self.params = {"l": 1, "sigma_f": 1}
        self.optimize = optimize

    def fit(self, X, y):
        # store train data
        self.train_X = np.asarray(X)
        self.train_y = np.asarray(y)
        self.is_fit = True

    def predict(self, X):
        if not self.is_fit:
            print("GPR Model not fit yet.")
            return

        X = np.asarray(X)
        Kff = self.kernel(self.train_X, self.train_X)  # (N, N)
        Kyy = self.kernel(X, X)  # (k, k)
        Kfy = self.kernel(self.train_X, X)  # (N, k)
        Kff_inv = np.linalg.inv(Kff + 1e-8 * np.eye(len(self.train_X)))  # (N, N)
        
        mu = Kfy.T.dot(Kff_inv).dot(self.train_y)
        cov = Kyy - Kfy.T.dot(Kff_inv).dot(Kfy)
        return mu, cov

    def kernel(self, x1, x2):
        dist_matrix = np.sum(x1**2, 1).reshape(-1, 1) + np.sum(x2**2, 1) - 2 * np.dot(x1, x2.T)
        return self.params["sigma_f"] ** 2 * np.exp(-0.5 / self.params["l"] ** 2 * dist_matrix)
def y(x, noise_sigma=0.0):
    x = np.asarray(x)
    y = np.cos(x) + np.random.normal(0, noise_sigma, size=x.shape)
    return y.tolist()

train_X = np.array(data_x).reshape(-1, 1)
train_y = np.array(data_n).reshape(-1, 1)
test_X = np.arange(-2.5, 2.5, 0.01).reshape(-1, 1)

gpr = GPR()
gpr.fit(train_X, train_y)
mu, cov = gpr.predict(test_X)
test_y = mu.ravel()
uncertainty = 1.96 * np.sqrt(np.diag(cov))
plt.figure()
plt.title("l=%.2f sigma_f=%.2f" % (gpr.params["l"], gpr.params["sigma_f"]))
plt.fill_between(test_X.ravel(), test_y + uncertainty, test_y - uncertainty, alpha=0.1)
plt.plot(test_X, test_y, label="回归曲线")
plt.scatter(train_X, train_y, label="带噪声的训练集", c="red",s=1)
plt.legend()
plt.show()