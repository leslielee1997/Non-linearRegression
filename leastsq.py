#最小二乘法
import numpy as np
import matplotlib.pyplot as plt
import openpyxl as xl
from scipy.optimize import leastsq
plt.rcParams['font.sans-serif']=['SimHei'] 
plt.rcParams['axes.unicode_minus']=False   
Filename = 'F:\code\meachine-learning\data.xlsx'
Sheet = 'Sheet1'
data = xl.load_workbook(Filename)[Sheet]
#读取数据集
def getColValues(data,column):
    rows = data.max_row
    columndata=[]
    for i in range(1,rows+1):
        cellvalue = data.cell(row=i,column=column).value
        columndata.append(cellvalue)
    return columndata
data_x = np.array(getColValues(data,1))
data_y = np.array(getColValues(data,2))
data_n = np.array(getColValues(data,3))
# 多项式函数
def fit_func(p, x):
    f = np.poly1d(p) 
    ret = f(x)
    return ret
# 计算残差
def residuals_func(p, x, y):
    ret = fit_func(p, x) - y
    return ret
#最小二乘法，M为多项式次数
def fitting(M=0):
    p_init = np.random.rand(M + 1)
    p_lsq = leastsq(residuals_func, p_init, args=(data_x, data_n))
    print('拟合系数为:', p_lsq[0])
    return p_lsq[0]
for i in range(1,4,1):
    info = '第'+ str(i) + '次拟合，多项式次数为' + str(i)
    print(info)
    plt.plot(data_x, data_y,'r-', linewidth=2, label=u'真实数据')
    plt.scatter(data_x, data_n, label='带噪声数据',s=5)
    plt.plot(data_x, fit_func(fitting(i), data_x), 'g-', linewidth=2, label=u'拟合数据')
    plt.legend()
    plt.show()