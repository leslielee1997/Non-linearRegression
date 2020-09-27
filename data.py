#输出数据集为xlsx文件,数据集为(x,y) y=x^2+sin(x)
import numpy as np
import openpyxl as xl
import random
Filename = 'F:\code\meachine-learning\data.xlsx'
Sheet = 'Sheet1'
home_work = xl.load_workbook(Filename)
hw = home_work[Sheet]
for i in range(1,100,1):
    x = (i*0.05-2.5)
    y = x*x+np.sin(x)
    hw.cell(row = i, column = 1).value = x 
    hw.cell(row = i, column = 2).value = y
    hw.cell(row = i, column = 3).value = y
x = np.linspace(1,99,9)
for i in x:
    hw.cell(row=int(np.ceil(i)),column = 3).value = hw.cell(row=int(np.ceil(i)),column = 2).value+np.random.normal(0, 0.3)
home_work.save(Filename)